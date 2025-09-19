# insertar_columna.py
import os
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Alignment, Protection, Side
import pandas as pd
from typing import List, Tuple, Optional, Union


def crear_borde_estilo(grosor: str = 'thin') -> Border:
    """Crea un estilo de borde consistente"""
    lado = Side(border_style=grosor, color='000000')
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def aplicar_bordes_tabla(sheet, fila_inicio: int, fila_fin: int, col_inicio: int, col_fin: int):
    """Aplica bordes a toda la tabla procesada"""
    try:
        # Estilo de borde para celdas normales
        borde_normal = crear_borde_estilo('thin')

        # Estilo de borde más grueso para el contorno de la tabla
        borde_contorno = crear_borde_estilo('medium')

        # Aplicar bordes a todas las celdas de la tabla
        for row in range(fila_inicio, fila_fin + 1):
            for col in range(col_inicio, col_fin + 1):
                cell = sheet.cell(row=row, column=col)

                # Aplicar borde normal a todas las celdas
                cell.border = borde_normal

                # Borde grueso en los contornos de la tabla
                if row == fila_inicio or row == fila_fin or col == col_inicio or col == col_fin:
                    cell.border = borde_contorno

        print(f"✅ Bordes aplicados a tabla: filas {fila_inicio}-{fila_fin}, columnas {col_inicio}-{col_fin}")

    except Exception as e:
        print(f"⚠️  Error al aplicar bordes: {str(e)}")


def aplicar_estilo_cabeceras(sheet, fila: int, col_inicio: int, col_fin: int):
    """Aplica estilo especial a las cabeceras de la tabla"""
    try:
        # Estilo para cabeceras
        fill_cabecera = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        font_cabecera = Font(bold=True, color="000000", size=11)
        alignment_cabecera = Alignment(horizontal="center", vertical="center", wrap_text=True)
        borde_cabecera = crear_borde_estilo('medium')

        for col in range(col_inicio, col_fin + 1):
            cell = sheet.cell(row=fila, column=col)
            cell.fill = fill_cabecera
            cell.font = font_cabecera
            cell.alignment = alignment_cabecera
            cell.border = borde_cabecera

        print(f"✅ Estilo aplicado a cabeceras en fila {fila}")

    except Exception as e:
        print(f"⚠️  Error al aplicar estilo a cabeceras: {str(e)}")


def ajustar_ancho_columnas(sheet):
    """Ajusta automáticamente el ancho de las columnas al contenido"""
    try:
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        # Calcular longitud del contenido
                        if isinstance(cell.value, datetime):
                            length = 10  # Longitud fija para fechas
                        else:
                            length = len(str(cell.value))

                        if length > max_length:
                            max_length = length
                except:
                    pass

            # Ajustar ancho con un poco de margen
            adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
            sheet.column_dimensions[column_letter].width = adjusted_width

        print("✅ Ancho de columnas ajustado automáticamente")

    except Exception as e:
        print(f"⚠️  Error al ajustar ancho de columnas: {str(e)}")


def eliminar_formatos(sheet) -> bool:
    """Elimina todos los formatos de la hoja para evitar errores"""
    try:
        # Eliminar merged cells primero
        merged_ranges = list(sheet.merged_cells.ranges)
        for merged_range in merged_ranges:
            sheet.unmerge_cells(str(merged_range))

        # Eliminar formatos solo de celdas con datos
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.font = Font()
                    cell.fill = PatternFill()
                    cell.border = Border()
                    cell.alignment = Alignment()
                    cell.number_format = 'General'

        print("✅ Formatos eliminados correctamente")
        return True

    except Exception as e:
        print(f"⚠️  Advertencia al eliminar formatos: {str(e)}")
        return False


def convertir_a_fecha_dd_mm_yyyy(valor) -> Union[datetime, None]:
    """Convierte un valor a datetime con formato dd/mm/yyyy"""
    if valor is None:
        return None

    if isinstance(valor, datetime):
        return valor

    if isinstance(valor, str):
        valor = valor.strip()
        if not valor:
            return None

        formatos_intentar = [
            '%d/%m/%Y', '%d/%m/%y', '%d-%m-%Y', '%d-%m-%y',
            '%d.%m.%Y', '%d.%m.%y', '%Y-%m-%d'
        ]

        for formato in formatos_intentar:
            try:
                fecha = datetime.strptime(valor, formato)
                if formato.endswith('%y') and fecha.year < 100:
                    if fecha.year < 50:
                        fecha = fecha.replace(year=fecha.year + 2000)
                    else:
                        fecha = fecha.replace(year=fecha.year + 1900)
                return fecha
            except ValueError:
                continue

    if isinstance(valor, (int, float)):
        try:
            if valor >= 1:
                fecha_base = datetime(1899, 12, 30)
                return fecha_base + pd.Timedelta(days=valor)
        except:
            pass

    return None


def formatear_fecha_dd_mm_yyyy(fecha: datetime) -> str:
    """Formatea datetime a string dd/mm/yyyy"""
    return fecha.strftime('%d/%m/%Y')


def aplicar_formato_fecha_excel(sheet, columna: int, desde_fila: int):
    """Aplica formato de fecha dd/mm/yyyy a la columna especificada"""
    for row in range(desde_fila, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=columna)
        if isinstance(cell.value, datetime):
            cell.number_format = 'DD/MM/YYYY'
        elif isinstance(cell.value, str) and re.match(r'\d{1,2}/\d{1,2}/\d{4}', cell.value):
            # Ya está en formato string dd/mm/yyyy
            pass


def procesar_excel(file_path: str) -> Tuple[bool, Optional[str], int]:
    """Función para procesar el archivo Excel con bordes y formato profesional"""
    try:
        # Cargar el workbook
        wb = load_workbook(filename=file_path, data_only=True)
        sheet = wb.active

        # 1. ELIMINAR FORMATOS primero para evitar errores
        eliminar_formatos(sheet)
        print("✅ Formatos eliminados")

        # 2. Insertar una columna en la posición A
        sheet.insert_cols(1)
        print("✅ Columna A insertada")

        # Encontrar la última fila y columna con datos reales
        last_row = 0
        last_column = 0

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    last_row = max(last_row, cell.row)
                    last_column = max(last_column, cell.column)

        print(f"📊 Filas: {last_row}, Columnas: {last_column}")

        # 3. Encontrar filas que comienzan con "6" y tienen más de 2 dígitos
        pattern_rows = []
        for row in range(1, last_row + 1):
            cell_value = sheet.cell(row=row, column=2).value
            if cell_value and isinstance(cell_value, str):
                cell_value = str(cell_value).strip()
                if cell_value.startswith('6'):
                    match = re.match(r'^(\d+)', cell_value)
                    if match and len(match.group(1)) > 2:
                        pattern_rows.append(row)

        print(f"🔍 Patrones encontrados: {len(pattern_rows)}")

        # 4. Copiar valores a la columna A para los patrones encontrados
        for i, current_pattern_row in enumerate(pattern_rows):
            next_pattern_row = pattern_rows[i + 1] if i < len(pattern_rows) - 1 else last_row + 1
            pattern_value = sheet.cell(row=current_pattern_row, column=2).value

            for row_num in range(current_pattern_row, next_pattern_row):
                sheet.cell(row=row_num, column=1).value = pattern_value

        print("✅ Valores copiados a columna A")

        # 5. ELIMINAR COLUMNAS K, L, M (columnas 11, 12, 13)
        columns_to_delete = [col for col in [11, 12, 13] if col <= last_column]

        for col in sorted(columns_to_delete, reverse=True):
            sheet.delete_cols(col)
            print(f"✅ Columna {get_column_letter(col)} eliminada")

        # Actualizar última columna
        last_column = sheet.max_column
        print(f"📊 Columnas después de eliminar K,L,M: {last_column}")

        # 6. RESTAR columna I - columna J y resultado en columna I
        if last_column >= 10:
            for row in range(7, last_row + 1):
                try:
                    valor_i = sheet.cell(row=row, column=9).value
                    valor_j = sheet.cell(row=row, column=10).value

                    try:
                        num_i = float(valor_i) if valor_i not in [None, ''] else 0
                    except (ValueError, TypeError):
                        num_i = 0

                    try:
                        num_j = float(valor_j) if valor_j not in [None, ''] else 0
                    except (ValueError, TypeError):
                        num_j = 0

                    sheet.cell(row=row, column=9).value = num_i - num_j

                except Exception as e:
                    print(f"⚠️  Error en fila {row}: {str(e)}")
                    continue

            print("✅ Resta I - J completada")

        # 7. AGREGAR CABECERAS en fila 6
        if last_row < 6:
            for _ in range(6 - last_row):
                sheet.insert_rows(last_row + 1)

        cabeceras = {
            1: "Cta",
            2: "Nro",
            3: "Suc - Tipo - Nro",
            4: "Fecha",
            5: "Org.",
            6: "Nro CPago - Tipo/Serie/ Numero/Fecha de Emision",
            7: "Glosa / Proveedor",
            8: "CC",
            9: "Debe"
        }

        for col_num, header_text in cabeceras.items():
            if col_num <= last_column:
                sheet.cell(row=6, column=col_num).value = header_text

        print("✅ Cabeceras agregadas en fila 6")

        # 8. PROCESAR FECHAS - Convertir y formatear a dd/mm/yyyy
        filas_a_eliminar = []
        filas_con_fecha = []

        for row in range(7, last_row + 1):
            fecha_valor = sheet.cell(row=row, column=4).value
            fecha_convertida = convertir_a_fecha_dd_mm_yyyy(fecha_valor)

            if fecha_convertida:
                fecha_formateada = formatear_fecha_dd_mm_yyyy(fecha_convertida)

                fila_datos = []
                for col in range(1, last_column + 1):
                    if col == 4:
                        fila_datos.append(fecha_formateada)
                    else:
                        fila_datos.append(sheet.cell(row=row, column=col).value)

                filas_con_fecha.append((fecha_convertida, fila_datos))
            else:
                filas_a_eliminar.append(row)

        # Eliminar filas sin fecha válida
        for row in sorted(filas_a_eliminar, reverse=True):
            sheet.delete_rows(row)

        print(f"✅ Filas sin fecha válida eliminadas: {len(filas_a_eliminar)}")
        print(f"✅ Filas con fecha válida: {len(filas_con_fecha)}")

        # 9. ORDENAR por fecha y ESCRIBIR DATOS
        if filas_con_fecha:
            filas_con_fecha.sort(key=lambda x: x[0])

            # Limpiar datos existentes desde fila 7
            for row in range(7, sheet.max_row + 1):
                for col in range(1, last_column + 1):
                    sheet.cell(row=row, column=col).value = None

            # Escribir datos ordenados
            for idx, (fecha_original, fila_datos) in enumerate(filas_con_fecha, 7):
                for col_idx, valor in enumerate(fila_datos, 1):
                    if col_idx <= last_column:
                        sheet.cell(row=idx, column=col_idx).value = valor

            # Aplicar formato de fecha Excel
            aplicar_formato_fecha_excel(sheet, 4, 7)

            print("✅ Datos ordenados por fecha y formateados a dd/mm/yyyy")

            # 10. APLICAR BORDES Y ESTILOS A LA TABLA
            fila_inicio_tabla = 6  # Cabeceras
            fila_fin_tabla = 6 + len(filas_con_fecha)  # Última fila con datos
            col_inicio_tabla = 1  # Columna A
            col_fin_tabla = last_column  # Última columna

            # Aplicar bordes a toda la tabla
            aplicar_bordes_tabla(sheet, fila_inicio_tabla, fila_fin_tabla, col_inicio_tabla, col_fin_tabla)

            # Aplicar estilo especial a las cabeceras
            aplicar_estilo_cabeceras(sheet, 6, col_inicio_tabla, col_fin_tabla)

            # Ajustar automáticamente el ancho de columnas
            ajustar_ancho_columnas(sheet)

        else:
            print("⚠️  No hay filas con fechas válidas para ordenar")

        # 11. Guardar el archivo procesado
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre_base = os.path.splitext(os.path.basename(file_path))[0]
        nuevo_nombre = f"procesado_{timestamp}_{nombre_base}.xlsx"
        nuevo_path = os.path.join(os.path.dirname(file_path), nuevo_nombre)

        wb.save(nuevo_path)
        print(f"💾 Archivo guardado como: {nuevo_nombre}")

        return True, nuevo_nombre, len(pattern_rows)

    except Exception as e:
        print(f"❌ Error crítico en procesar_excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return False, None, 0


def validar_procesamiento(file_path: str):
    """Valida que el archivo procesado tenga el formato correcto"""
    try:
        wb = load_workbook(file_path)
        sheet = wb.active

        # Verificar bordes en la primera celda de la tabla
        celda_prueba = sheet.cell(row=6, column=1)
        tiene_bordes = celda_prueba.border.left.style is not None

        print(f"✅ Validación: Tabla con bordes → {'SÍ' if tiene_bordes else 'NO'}")
        return tiene_bordes

    except Exception as e:
        print(f"❌ Error en validación: {str(e)}")
        return False


# Ejemplo de uso
if __name__ == "__main__":
    archivo = "tu_archivo.xlsx"  # Reemplaza con tu ruta

    print("🔄 Iniciando procesamiento con bordes...")
    success, nombre_archivo, patrones = procesar_excel(archivo)

    if success:
        print(f"🎉 Procesamiento completado. Patrones encontrados: {patrones}")
        print("🔍 Validando estilo de la tabla...")
        if validar_procesamiento(nombre_archivo):
            print("✅ Tabla procesada con bordes y estilo profesional")
        else:
            print("⚠️  La tabla no tiene el estilo esperado")
    else:
        print("❌ Error en el procesamiento")
