import pandas as pd
from openpyxl import load_workbook
import os
import shutil
import re
from datetime import datetime
import hashlib

# Configuración de la contraseña (debe ser la misma)
PASSWORD_HASH = "c8a6ed3ac08087cc037c2fc7846a7f95976b8f5bfbaf2d9540cf89b74452b034"


def verificar_password(password):
    """Verifica la contraseña"""
    if not password:
        return False
    hash_input = hashlib.sha256(password.encode()).hexdigest()
    return hash_input == PASSWORD_HASH


def obtener_hojas_analisis(origen_path):
    """Obtiene todas las hojas que comienzan con 'Analisis'"""
    try:
        xl = pd.ExcelFile(origen_path)
        hojas_analisis = [sheet for sheet in xl.sheet_names if sheet.startswith('Analisis')]

        if not hojas_analisis:
            raise ValueError("No se encontraron hojas que comiencen con 'Analisis'")

        return hojas_analisis
    except Exception as e:
        raise ValueError(f"Error al leer las hojas del archivo: {str(e)}")


def limpiar_glosa_proveedor(texto):
    """Limpia el texto de Glosa/Proveedor removiendo guiones al final"""
    if pd.isna(texto):
        return texto

    texto = str(texto).strip()

    patrones = [
        (r'ESTABLE\s*-\s*$', 'ESTABLE'),
        (r'CONTRATADO\s*-\s*$', 'CONTRATADO'),
        (r'20530\s*-\s*$', '20530'),
        (r'\s*-\s*$', ''),
    ]

    for patron, reemplazo in patrones:
        texto = re.sub(patron, reemplazo, texto)

    return texto.strip()


def formatear_fecha(fecha):
    """Asegura que las fechas tengan el formato correcto"""
    if pd.isna(fecha):
        return fecha

    if isinstance(fecha, datetime) or hasattr(fecha, 'strftime'):
        return fecha

    try:
        if isinstance(fecha, str):
            formatos = ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%m-%d-%Y']
            for formato in formatos:
                try:
                    return datetime.strptime(fecha, formato)
                except ValueError:
                    continue
    except:
        pass

    return fecha


def formatear_numero(valor):
    """Convierte el valor a formato numérico para Excel"""
    if pd.isna(valor):
        return valor

    if isinstance(valor, (int, float)):
        return float(valor)

    try:
        if isinstance(valor, str):
            valor_limpio = re.sub(r'[^\d\.\-]', '', valor.strip())
            if valor_limpio:
                return float(valor_limpio)
    except:
        pass

    return valor


def procesar_transferencia(origen_path, destino_path, hoja_seleccionada, password):
    """Función principal para transferir datos"""
    try:
        # Verificar contraseña
        if not verificar_password(password):
            return False, "Contraseña incorrecta", None, None

        # Configuración de mapeo de columnas
        mapeo_columnas = [
            {'origen': 'Cta', 'destino': 'CTA', 'col_origen': None, 'col_destino': None},
            {'origen': 'Suc - Tipo - Nro', 'destino': 'Suc - Tipo - Nro', 'col_origen': 3, 'col_destino': 5},
            {'origen': 'Fecha', 'destino': 'FECHA', 'col_origen': 4, 'col_destino': 6},
            {'origen': 'Glosa / Proveedor', 'destino': 'Glosa / Proveedor', 'col_origen': 7, 'col_destino': 9},
            {'origen': 'CC', 'destino': 'CC', 'col_origen': 8, 'col_destino': 10},
            {'origen': 'Debe', 'destino': 'Debe', 'col_origen': 9, 'col_destino': 12, 'formato': 'numero'}
        ]

        # Leer datos del archivo ORIGEN
        df_origen = pd.read_excel(origen_path, sheet_name=hoja_seleccionada, header=5)

        # Verificar columnas
        columnas_faltantes = []
        for mapeo in mapeo_columnas:
            if mapeo['origen'] not in df_origen.columns:
                columnas_faltantes.append(mapeo['origen'])

        if columnas_faltantes:
            raise ValueError(f"Columnas no encontradas en origen: {columnas_faltantes}")

        # Limpiar y formatear datos
        df_origen['Glosa / Proveedor'] = df_origen['Glosa / Proveedor'].apply(limpiar_glosa_proveedor)
        df_origen['Fecha'] = df_origen['Fecha'].apply(formatear_fecha)
        df_origen['Debe'] = df_origen['Debe'].apply(formatear_numero)

        # Cargar archivo DESTINO
        libro_destino = load_workbook(destino_path)

        if 'BD6' not in libro_destino.sheetnames:
            raise ValueError("No se encontró la hoja 'BD6' en el archivo destino")

        hoja_destino = libro_destino['BD6']

        # Configurar números de columna
        for mapeo in mapeo_columnas:
            if mapeo['col_destino'] is None:
                for col in range(1, hoja_destino.max_column + 1):
                    if hoja_destino.cell(row=5, column=col).value == mapeo['destino']:
                        mapeo['col_destino'] = col
                        break

        # Verificar columnas destino
        columnas_destino_faltantes = []
        for mapeo in mapeo_columnas:
            if mapeo['col_destino'] is None:
                columnas_destino_faltantes.append(mapeo['destino'])

        if columnas_destino_faltantes:
            raise ValueError(f"Columnas no encontradas en destino: {columnas_destino_faltantes}")

        # Limpiar columnas destino
        for mapeo in mapeo_columnas:
            col_dest = mapeo['col_destino']
            for row in range(6, hoja_destino.max_row + 1):
                hoja_destino.cell(row=row, column=col_dest).value = None

        # Transferir datos
        filas_transferidas = 0
        for idx, fila in df_origen.iterrows():
            fila_destino = 6 + idx
            tiene_datos = any(pd.notna(fila[mapeo['origen']]) for mapeo in mapeo_columnas)

            if tiene_datos:
                if fila_destino > hoja_destino.max_row:
                    nueva_fila = [None] * hoja_destino.max_column
                    hoja_destino.append(nueva_fila)

                for mapeo in mapeo_columnas:
                    valor = fila[mapeo['origen']]
                    if pd.notna(valor):
                        if mapeo.get('formato') == 'numero' and isinstance(valor, (int, float)):
                            hoja_destino.cell(row=fila_destino, column=mapeo['col_destino']).value = float(valor)
                        else:
                            hoja_destino.cell(row=fila_destino, column=mapeo['col_destino']).value = valor

                filas_transferidas += 1

        # Crear backup
        nombre_base = os.path.splitext(destino_path)[0]
        extension = os.path.splitext(destino_path)[1]
        backup_path = f"{nombre_base}_backup{extension}"
        shutil.copy2(destino_path, backup_path)

        # Guardar cambios
        libro_destino.save(destino_path)

        # Preparar resumen
        resumen = {
            'archivo_origen': os.path.basename(origen_path),
            'hoja_origen': hoja_seleccionada,
            'archivo_destino': os.path.basename(destino_path),
            'filas_transferidas': filas_transferidas,
            'columnas_transferidas': [mapeo['origen'] for mapeo in mapeo_columnas],
            'backup_path': backup_path
        }

        return True, "Transferencia completada exitosamente", resumen, destino_path

    except Exception as e:
        return False, f"Error durante la transferencia: {str(e)}", None, None
